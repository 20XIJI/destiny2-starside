#!/usr/bin/env python3
"""把 tools/sheet-grab.user.js 抓下来的 JSON 还原成 xlsx，便于核对与二次编辑。

用法：
    python3 tools/json2xlsx.py <抓取的.json> [输出.xlsx]

还原的是**替换后**的表：文本按抓取时页面上渲染的样子写入（术语替换已生效）。
Google 官方的 export?format=xlsx 给的是替换前的原文，且不少表禁用了导出，所以
这条路不能替代它。

图优先用 JSON 顶层 images 里内联的 base64，全程不联网。只有 URL 的老 JSON 才
现下载，而 URL 是有寿命的：隔天再跑可能整批取不到，届时重新抓一份。

图片取不到不吞：逐条打印到 stderr，并以非零退出码结束——静默少几张图，
在几百行的表里眼睛查不出来。
"""

import argparse
import base64
import hashlib
import json
import os
import sys
import tempfile
import urllib.request


PX_PER_CHAR = 7.0        # 列宽单位是「字符」，约 7 px 一个
PT_PER_PX = 0.75         # 行高单位是「磅」
MAX_COL_CHARS = 64       # 说明列动辄上百字，不封顶会把整表拉成一条线
MIN_ROW_PT = 15.0


def die(msg):
    print('转换中止：' + msg, file=sys.stderr)
    raise SystemExit(1)


def fetch(url, cache, inline):
    """先用 JSON 里内联的 base64，没有才下载。按 URL 的 md5 落盘缓存，同一张图
    在表里出现多次只解一次。"""
    key = hashlib.md5(url.encode()).hexdigest()
    for ext in ('.png', '.jpg'):
        p = os.path.join(cache, key + ext)
        if os.path.exists(p):
            return p
    if url in inline:
        blob = base64.b64decode(inline[url].split(',', 1)[1])
    else:                                       # 老 JSON 只有 URL
        with urllib.request.urlopen(url, timeout=30) as r:
            blob = r.read()
    ext = '.png' if blob[:8] == b'\x89PNG\r\n\x1a\n' else '.jpg'
    p = os.path.join(cache, key + ext)
    with open(p, 'wb') as f:
        f.write(blob)
    return p


def build(data, cache, ws):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    inline = data.get('images') or {}
    widths, heights, failed, placed = {}, {}, [], 0
    for r, row in enumerate(data['rows'], start=1):
        for c, cell in enumerate(row, start=1):
            text = cell.get('t', '')
            if text:
                ws.cell(row=r, column=c, value=text).alignment = Alignment(
                    wrap_text=True, vertical='top')
                # 列宽按最长的那一行估，不按整段长度——整段会把宽度顶到封顶值
                longest = max((len(ln) for ln in text.split('\n')), default=0)
                widths[c] = max(widths.get(c, 0), min(longest, MAX_COL_CHARS))
                heights[r] = max(heights.get(r, 0), MIN_ROW_PT * (text.count('\n') + 1))
            for url in cell.get('img', []):
                try:
                    path = fetch(url, cache, inline)
                except Exception as e:                      # noqa: BLE001
                    failed.append('r%d c%d %s —— %s' % (r, c, url, e))
                    continue
                img = XLImage(path)
                ws.add_image(img, '%s%d' % (get_column_letter(c), r))
                widths[c] = max(widths.get(c, 0), img.width / PX_PER_CHAR)
                heights[r] = max(heights.get(r, 0), img.height * PT_PER_PX)
                placed += 1
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w + 2
    for r, h in heights.items():
        ws.row_dimensions[r].height = h
    return placed, failed


def main(argv):
    parser = argparse.ArgumentParser(description=__doc__, allow_abbrev=False)
    parser.add_argument('src', help='sheet-grab 导出的 JSON')
    parser.add_argument('out', nargs='?', help='输出路径，默认输入同名 .xlsx')
    parser.add_argument('--force', action='store_true', help='显式覆盖已有完整输出')
    args = parser.parse_args(argv[1:])
    src = args.src
    out = args.out or os.path.splitext(src)[0] + '.xlsx'
    if (os.path.realpath(src) == os.path.realpath(out)
            or (os.path.exists(src) and os.path.exists(out) and os.path.samefile(src, out))):
        die('输入与输出不能是同一个文件')
    if os.path.isdir(out):
        die('输出是目录，不能覆盖：' + out)
    if os.path.lexists(out) and not args.force:
        die('输出已存在：%s；另给输出文件名，或显式使用 --force 覆盖' % out)
    parent = os.path.dirname(os.path.abspath(out))
    if not os.path.isdir(parent):
        die('输出父目录不存在：' + parent)

    from openpyxl import Workbook

    with open(src, encoding='utf-8') as f:
        data = json.load(f)
    if 'rows' not in data:
        die('%s 里没有 rows，这不是 sheet-grab 抓出来的 JSON' % src)

    wb = Workbook()
    # 新建的 Workbook 一定有活动表，但 openpyxl 的类型标注写的是 Optional
    ws = wb.active if wb.active is not None else wb.create_sheet()
    ws.title = (data.get('title') or 'sheet')[:31]

    published = False
    with tempfile.TemporaryDirectory(dir=parent) as cache:
        placed, failed = build(data, cache, ws)
        candidate = os.path.join(cache, 'candidate.xlsx')
        try:
            wb.save(candidate)          # 图片在 save 时才读，缓存目录要活到这一步
            if failed and os.path.lexists(out):
                print('转换缺图，未覆盖已有输出：' + out, file=sys.stderr)
            elif args.force:
                os.replace(candidate, out)
                published = True
            else:
                os.link(candidate, out)  # 预检后出现的目标也不能覆盖
                published = True
        except OSError as e:
            die('未发布输出 %s：%s' % (out, e))

    cells = sum(1 for row in data['rows'] for cell in row if cell.get('t'))
    if not failed:
        print('%s —— %d 行，%d 格有字，%d 张图' % (out, len(data['rows']), cells, placed))
    elif published:
        print('部分结果（缺图）：' + out)
    if failed:
        print('这 %d 张图没下下来：' % len(failed), file=sys.stderr)
        for line in failed:
            print('  ' + line, file=sys.stderr)
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
